from __future__ import annotations

import argparse
import asyncio
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml
from ib_insync import IB, Contract, util
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# ── Logging setup ────────────────────────────────────────────────────────────

def setup_logging(cfg: dict):
    level = getattr(logging, cfg.get("level", "INFO").upper(), logging.INFO)
    handlers: list[logging.Handler] = [logging.StreamHandler(sys.stdout)]

    log_file = cfg.get("file")
    if log_file:
        date_prefix = datetime.now().strftime("%Y_%m_%d")
        daily_log = Path(log_file).parent / f"{date_prefix}_snapshot_bot.log"
        handlers.append(logging.FileHandler(daily_log, encoding="utf-8"))

    logging.basicConfig(
        level=level,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=handlers,
    )
    # Suppress all ib_insync internal logging — we handle errors via errorEvent
    for name in ("ib_insync", "ib_insync.wrapper", "ib_insync.client", "ib_insync.ib"):
        logging.getLogger(name).setLevel(logging.CRITICAL)


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


# ── IB helpers ───────────────────────────────────────────────────────────────

log = logging.getLogger(__name__)

IGNORED_ERRORS = {2104, 2106, 2107, 2158, 2119, 300, 10089, 2103}
WARNING_ERRORS = {1100, 1101, 1102, 321, 2151, 2137}


def on_error(reqId, errorCode, errorString, contract):
    if errorCode in IGNORED_ERRORS:
        return
    if errorCode in WARNING_ERRORS:
        log.warning("IBKR notice %d: %s", errorCode, errorString)
    else:
        log.error("IBKR error %d (reqId=%d): %s", errorCode, reqId, errorString)


def get_last_price(ib: IB, contract: Contract) -> Optional[float]:
    ticker = ib.reqMktData(contract, "", snapshot=True, regulatorySnapshot=False)
    result = None
    for _ in range(10):
        ib.sleep(0.5)
        price = ticker.last
        if price and price == price and price > 0:
            result = price
            break
        price = ticker.close
        if price and price == price and price > 0:
            result = price
            break
    try:
        ib.cancelMktData(contract)
    except Exception:
        pass
    return result


def get_company_name(ib: IB, contract: Contract) -> str:
    try:
        details = ib.reqContractDetails(contract)
        if details:
            return details[0].longName
    except Exception:
        pass
    return ""


def smart_contract(conid: int, symbol: str, sec_type: str, currency: str) -> Contract:
    c = Contract()
    c.conId = conid
    c.symbol = symbol
    c.secType = sec_type
    c.exchange = "SMART"
    c.currency = currency
    return c


# ── Snapshot logic ───────────────────────────────────────────────────────────

COLUMNS = [
    "Date", "Time", "Symbol", "Company Name", "Quantity",
    "Entry Price", "Current Price", "Take Profit Price", "Stop Loss Price",
    "Trailing Active", "Trailing %", "PnL USD", "PnL %", "Behaviour",
]


async def run_snapshot(ib: IB, output_dir: str):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    log.info("=== Portfolio Snapshot started ===")

    ib.reqMarketDataType(4)
    ib.reqPositions()
    ib.reqAllOpenOrders()
    await asyncio.sleep(3)

    positions = [p for p in ib.positions() if p.position != 0]
    open_trades = ib.openTrades()

    # Index orders by conId
    orders_by_conid: dict[int, list] = {}
    for t in open_trades:
        cid = t.contract.conId
        orders_by_conid.setdefault(cid, []).append(t)

    log.info("Found %d open position(s).", len(positions))

    rows = []
    for pos in positions:
        c = pos.contract
        conid = c.conId
        symbol = c.symbol
        qty = pos.position
        entry_price = pos.avgCost
        is_long = qty > 0

        contract = smart_contract(conid, symbol, c.secType, c.currency)

        # Get current price
        current_price = get_last_price(ib, contract)
        log.info("  %s: qty=%s entry=%.2f current=%s",
                 symbol, qty, entry_price, f"{current_price:.2f}" if current_price else "N/A")

        # Get company name
        company_name = get_company_name(ib, contract)

        # Detect TP / SL / TRAIL from open orders
        tp_price = None
        sl_price = None
        trailing_active = False
        trailing_pct = None

        for t in orders_by_conid.get(conid, []):
            otype = t.order.orderType
            status = t.orderStatus.status
            if status in ("Cancelled", "Inactive", "Filled"):
                continue
            if otype == "LMT":
                tp_price = t.order.lmtPrice
            elif otype == "STP":
                sl_price = t.order.auxPrice
            elif otype == "TRAIL":
                trailing_active = True
                trailing_pct = t.order.trailingPercent

        # PnL calculation
        pnl_usd = None
        pnl_pct = None
        if current_price and entry_price > 0:
            pnl_usd = (current_price - entry_price) * qty
            pnl_pct = ((current_price - entry_price) / entry_price) * 100

        # Behaviour
        if trailing_active:
            behaviour = "Trailing Stop"
        elif tp_price and sl_price:
            behaviour = "TP/SL Protected"
        elif tp_price:
            behaviour = "TP Only"
        elif sl_price:
            behaviour = "SL Only"
        else:
            behaviour = "Unprotected"

        rows.append({
            "Date": date_str,
            "Time": time_str,
            "Symbol": symbol,
            "Company Name": company_name,
            "Quantity": qty,
            "Entry Price": entry_price,
            "Current Price": current_price,
            "Take Profit Price": tp_price,
            "Stop Loss Price": sl_price,
            "Trailing Active": "Yes" if trailing_active else "No",
            "Trailing %": trailing_pct,
            "PnL USD": pnl_usd,
            "PnL %": pnl_pct,
            "Behaviour": behaviour,
        })

    # Ensure output directory exists
    out = Path(output_dir) / "Snapshots"
    out.mkdir(parents=True, exist_ok=True)

    base = f"Portfolio_Snapshot_{now.strftime('%Y-%m-%d_%H-%M-%S')}"

    # Generate XLSX
    xlsx_path = out / f"{base}.xlsx"
    _write_xlsx(xlsx_path, rows)
    log.info("XLSX saved: %s", xlsx_path)

    # Generate PDF
    pdf_path = out / f"{base}.pdf"
    _write_pdf(pdf_path, rows, date_str, time_str)
    log.info("PDF  saved: %s", pdf_path)

    log.info("=== Portfolio Snapshot complete ===")


# ── XLSX generation ──────────────────────────────────────────────────────────

def _write_xlsx(filepath: Path, rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Snapshot"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data
    money_fmt = '#,##0.00'
    pct_fmt = '0.00"%"'

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")

            # Number formatting
            if col_name in ("Entry Price", "Current Price", "Take Profit Price",
                            "Stop Loss Price", "PnL USD"):
                if val is not None:
                    cell.number_format = money_fmt
            elif col_name in ("PnL %", "Trailing %"):
                if val is not None:
                    cell.number_format = pct_fmt

            # Color PnL
            if col_name == "PnL USD" and val is not None:
                cell.font = Font(color="006100" if val >= 0 else "9C0006")
            elif col_name == "PnL %" and val is not None:
                cell.font = Font(color="006100" if val >= 0 else "9C0006")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto column width
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(ws.cell(row=1, column=col_idx).value))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    wb.save(filepath)


# ── PDF generation ───────────────────────────────────────────────────────────

def _fmt_val(col: str, val) -> str:
    if val is None:
        return "-"
    if col in ("Entry Price", "Current Price", "Take Profit Price",
               "Stop Loss Price", "PnL USD"):
        return f"{val:,.2f}"
    if col in ("PnL %", "Trailing %"):
        return f"{val:.2f}%"
    return str(val)


def _write_pdf(filepath: Path, rows: list[dict], date_str: str, time_str: str):
    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        "SnapshotTitle", parent=styles["Title"], fontSize=14, spaceAfter=6,
    )
    elements.append(Paragraph(f"Portfolio Snapshot — {date_str} {time_str}", title_style))
    elements.append(Spacer(1, 4 * mm))

    # Table data
    header = COLUMNS
    table_data = [header]

    for row in rows:
        table_data.append([_fmt_val(col, row.get(col)) for col in COLUMNS])

    if not rows:
        table_data.append(["No open positions"] + [""] * (len(COLUMNS) - 1))

    col_count = len(COLUMNS)
    avail_width = landscape(A4)[0] - 20 * mm
    col_width = avail_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    # Color PnL cells
    for row_idx, row in enumerate(rows, 1):
        for col_name in ("PnL USD", "PnL %"):
            col_idx = COLUMNS.index(col_name)
            val = row.get(col_name)
            if val is not None:
                color = colors.HexColor("#006100") if val >= 0 else colors.HexColor("#9C0006")
                style_cmds.append(("TEXTCOLOR", (col_idx, row_idx), (col_idx, row_idx), color))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)


# ── Main ─────────────────────────────────────────────────────────────────────

async def main(cfg: dict, output_dir: str):
    ic = cfg["ibkr"]
    client_id = ic.get("snapshot_client_id", 4)

    ib = IB()
    ib.errorEvent += on_error

    try:
        ib.connect(
            host=ic["host"],
            port=ic["port"],
            clientId=client_id,
            timeout=20,
            readonly=True,
        )
        log.info("Connected to IB Gateway at %s:%s (clientId=%s)",
                 ic["host"], ic["port"], client_id)
    except Exception as exc:
        log.error("Connection failed: %s", exc)
        sys.exit(1)

    try:
        await run_snapshot(ib, output_dir)
    except Exception as exc:
        log.error("Fatal error: %s", exc, exc_info=True)
    finally:
        ib.disconnect()
        log.info("Disconnected.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="IBKR Portfolio Snapshot Bot")
    parser.add_argument("--config", default="config.yaml", help="Path to YAML config")
    parser.add_argument("--output", default=".", help="Base directory (files saved to <output>/Snapshots/)")
    args = parser.parse_args()

    if not Path(args.config).exists():
        print(f"Config file not found: {args.config}", file=sys.stderr)
        sys.exit(1)

    cfg = load_config(args.config)
    setup_logging(cfg.get("logging", {}))
    util.patchAsyncio()

    try:
        asyncio.run(main(cfg, args.output))
    except KeyboardInterrupt:
        log.info("Stopped by user.")
